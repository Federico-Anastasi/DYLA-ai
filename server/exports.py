"""Builds the exports (xlsx / drawio / html) from the JSON documents
(estimate.json, data_model.json) stored under projects/{name}/. dev_tasks.xlsx is
also built from estimate.json (epics[].tasks[].dev_tasks[]) — there is no separate
dev_tasks.json.

The visual styling follows the legacy templates in templates/generate_*.py (colors,
cell merging, fonts). The difference is where the data comes from: validated JSON
documents instead of DATA_* variables filled in by hand.
"""
from __future__ import annotations

import io
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import lanes as planner
from .documents import load_doc

# ── shared styles (openpyxl) ────────────────────────────────────────────

_THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_EPIC_FILL = PatternFill("solid", fgColor="D6E4F0")
_EPIC_FONT = Font(name="Arial", bold=True, size=10)
_E2E_FILL = PatternFill("solid", fgColor="F2F2F2")
_E2E_FONT = Font(name="Arial", italic=True, size=10)
_TASK_FILL = PatternFill("solid", fgColor="F2F2F2")
_TASK_FONT = Font(name="Arial", bold=True, size=10)
_CONT_FILL = PatternFill("solid", fgColor="D9D9D9")
_CONT_FONT = Font(name="Arial", bold=True, size=10)
_GRAND_FILL = PatternFill("solid", fgColor="C6EFCE")
_GRAND_FONT = Font(name="Arial", bold=True, size=12)
_NORMAL_FONT = Font(name="Arial", size=10)
_NOTE_FONT = Font(name="Arial", size=10, color="555555")
_SECTION_FILL = PatternFill("solid", fgColor="4472C4")
_SECTION_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Outcome colors for test_plan_xlsx: "ok" reuses the same green as _GRAND_FILL so the
# two workbooks read consistently; ko/blocked use Excel's standard red/yellow from
# conditional formatting.
_OUTCOME_OK_FILL = PatternFill("solid", fgColor="C6EFCE")
_OUTCOME_KO_FILL = PatternFill("solid", fgColor="FFC7CE")
_OUTCOME_BLOCKED_FILL = PatternFill("solid", fgColor="FFEB9C")
_OUTCOME_NEUTRAL_FILL = PatternFill("solid", fgColor="F2F2F2")
_OUTCOME_FILLS = {
    "ok": _OUTCOME_OK_FILL, "ko": _OUTCOME_KO_FILL,
    "blocked": _OUTCOME_BLOCKED_FILL, "to_run": _OUTCOME_NEUTRAL_FILL,
}
_OUTCOME_LABEL = {"ok": "OK", "ko": "KO", "blocked": "Blocked", "to_run": "To run"}


class DocNotFound(FileNotFoundError):
    """Raised when the JSON document an export is built from does not exist."""


def _require(project: str, doc: str) -> dict:
    data = load_doc(project, doc)
    if data is None:
        raise DocNotFound(f"{doc}.json not found for project '{project}'")
    return data


def _round_days(x: float) -> float:
    # Days are expressed with at most one decimal (see estimation_scale.md); rounding
    # to 1 decimal also keeps floating point artifacts out of the sheet
    # (e.g. 23.5 * 0.15 = 3.525 -> round(..., 2) can yield 3.52 instead of 3.5/3.53).
    return round(x, 1)


def _header_row(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN


def _section_header(ws, row, text, col_count):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    for col in range(1, col_count + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _SECTION_FILL
        c.border = _THIN


# ── estimate.xlsx ────────────────────────────────────────────────────────

def estimate_xlsx(project: str) -> bytes:
    data = _require(project, "estimate")
    meta = data["meta"]
    epics = data["epics"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 40

    _header_row(ws, ["Epic", "Task", "Days (Dev + Unit Test)", "Description"])

    row = 2
    total_days = 0.0
    epic_groups = []  # (start, end, name)

    for epic in epics:
        name = epic["name"]
        start = row
        for t in epic["tasks"]:
            ws.cell(row=row, column=2, value=t["task"])
            ws.cell(row=row, column=3, value=t["days"])
            ws.cell(row=row, column=4, value=t.get("description") or "")
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = _NORMAL_FONT
                cell.border = _THIN
                cell.alignment = _CENTER if col == 3 else _LEFT_WRAP
            if t.get("description"):
                ws.cell(row=row, column=4).font = _NOTE_FONT
            total_days += t["days"]
            row += 1
        if epic.get("e2e"):
            e2e = epic["e2e"]
            ws.cell(row=row, column=2, value=e2e["label"])
            ws.cell(row=row, column=3, value=e2e["days"])
            ws.cell(row=row, column=4, value="")
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = _E2E_FONT
                cell.fill = _E2E_FILL
                cell.border = _THIN
                cell.alignment = _CENTER if col == 3 else _LEFT_WRAP
            total_days += e2e["days"]
            row += 1
        if row == start:
            # An epic with no tasks and no E2E row (the schema allows both: "tasks" has no
            # minItems and "e2e" is optional) would otherwise leave `end < start` below: the
            # merge is skipped and the epic's own name lands on the SAME row as the next
            # epic's first task — whichever gets written there second wins, and the other
            # epic's name silently disappears from the file handed to the client. One
            # placeholder row keeps every epic on a row of its own.
            ws.cell(row=row, column=2, value="(no tasks yet)")
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value="")
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = _NOTE_FONT
                cell.border = _THIN
                cell.alignment = _CENTER if col == 3 else _LEFT_WRAP
            row += 1
        end = row - 1
        epic_groups.append((start, end, name))

    for start, end, name in epic_groups:
        if start < end:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        cell = ws.cell(row=start, column=1, value=name)
        cell.font = _EPIC_FONT
        cell.fill = _EPIC_FILL
        cell.alignment = _LEFT_WRAP
        cell.border = _THIN
        for r in range(start, end + 1):
            c = ws.cell(row=r, column=1)
            c.fill = _EPIC_FILL
            c.border = _THIN

    # Contingency
    contingency_pct = meta.get("contingency_pct", 0) or 0
    contingency_days = _round_days(total_days * contingency_pct / 100)
    ws.cell(row=row, column=1, value="Contingency")
    ws.cell(row=row, column=2, value=f"Buffer ~{contingency_pct:g}%")
    ws.cell(row=row, column=3, value=contingency_days)
    ws.cell(row=row, column=4, value="")
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.font = _CONT_FONT
        cell.fill = _CONT_FILL
        cell.border = _THIN
        cell.alignment = _CENTER if col == 3 else _LEFT_WRAP
    row += 1

    grand_days = _round_days(total_days + contingency_days)
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value="")
    ws.cell(row=row, column=3, value=grand_days)
    ws.cell(row=row, column=4, value="")
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.font = _GRAND_FONT
        cell.fill = _GRAND_FILL
        cell.border = _THIN
        cell.alignment = _CENTER if col == 3 else _LEFT_WRAP

    # Data Model sheet (when the document exists)
    dm = load_doc(project, "data_model")
    if dm:
        ws2 = wb.create_sheet("Data Model")
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 16
        ws2.column_dimensions["D"].width = 8
        ws2.column_dimensions["E"].width = 22
        ws2.column_dimensions["F"].width = 10
        ws2.column_dimensions["G"].width = 40

        r = 1
        _section_header(ws2, r, "Table fields", 7)
        r += 1
        for col, h in enumerate(["Table", "Field", "Type", "PK", "FK", "Nullable", "Notes"], 1):
            cell = ws2.cell(row=r, column=col, value=h)
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cell.border = _THIN
            cell.alignment = _CENTER
        r += 1
        for tbl in dm["tables"]:
            for f in tbl["fields"]:
                vals = [tbl["name"], f["name"], f["type"],
                        "PK" if f.get("pk") else "", f.get("fk") or "",
                        "" if f.get("nullable") is None else ("Yes" if f["nullable"] else "No"),
                        # schemas/data_model.schema.json calls the field "notes" (plural):
                        # this used to read "note" and always got None, so the column was
                        # empty for every field in every export.
                        f.get("notes") or ""]
                for col, val in enumerate(vals, 1):
                    cell = ws2.cell(row=r, column=col, value=val)
                    cell.font = _NORMAL_FONT
                    cell.border = _THIN
                    cell.alignment = _CENTER if col in (4,) else _LEFT_WRAP
                r += 1

        r += 1
        _section_header(ws2, r, "Relations", 4)
        r += 1
        for col, h in enumerate(["From", "To", "Type", "Label"], 1):
            cell = ws2.cell(row=r, column=col, value=h)
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cell.border = _THIN
            cell.alignment = _CENTER
        r += 1
        for rel in dm["relations"]:
            vals = [rel["from"], rel["to"], rel["type"], rel.get("label") or ""]
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=col, value=val)
                cell.font = _NORMAL_FONT
                cell.border = _THIN
                cell.alignment = _LEFT_WRAP
            r += 1

    # Assumptions and Considerations sheet
    assumptions = data.get("assumptions") or []
    considerations = data.get("considerations") or []
    if assumptions or considerations:
        ws3 = wb.create_sheet("Assumptions and Considerations")
        ws3.column_dimensions["A"].width = 8
        ws3.column_dimensions["B"].width = 100

        r = 1
        _section_header(ws3, r, "Assumptions", 2)
        r += 1
        for i, a in enumerate(assumptions, 1):
            ws3.cell(row=r, column=1, value=i).font = Font(name="Arial", bold=True, size=10)
            ws3.cell(row=r, column=1).alignment = _CENTER
            ws3.cell(row=r, column=1).border = _THIN
            cell = ws3.cell(row=r, column=2, value=a)
            cell.font = _NORMAL_FONT
            cell.alignment = _LEFT_TOP
            cell.border = _THIN
            r += 1

        r += 1
        _section_header(ws3, r, "General Considerations", 2)
        r += 1
        for i, c in enumerate(considerations, 1):
            ws3.cell(row=r, column=1, value=i).font = Font(name="Arial", bold=True, size=10)
            ws3.cell(row=r, column=1).alignment = _CENTER
            ws3.cell(row=r, column=1).border = _THIN
            cell = ws3.cell(row=r, column=2, value=c)
            cell.font = _NORMAL_FONT
            cell.alignment = _LEFT_TOP
            cell.border = _THIN
            r += 1

    # Open Questions sheet
    open_questions = data.get("open_questions") or []
    if open_questions:
        ws4 = wb.create_sheet("Open Questions")
        ws4.column_dimensions["A"].width = 5
        ws4.column_dimensions["B"].width = 20
        ws4.column_dimensions["C"].width = 60
        ws4.column_dimensions["D"].width = 20
        ws4.column_dimensions["E"].width = 12

        _header_row(ws4, ["#", "Area", "Question", "Estimated Impact", "Priority"])
        r = 2
        for i, q in enumerate(open_questions, 1):
            ws4.cell(row=r, column=1, value=i).font = _NORMAL_FONT
            ws4.cell(row=r, column=1).alignment = _CENTER
            ws4.cell(row=r, column=1).border = _THIN
            vals = [q.get("area", ""), q.get("question", ""), q.get("estimated_impact") or "", q.get("priority") or ""]
            for col, val in enumerate(vals, 2):
                cell = ws4.cell(row=r, column=col, value=val)
                cell.font = _NORMAL_FONT
                cell.alignment = _LEFT_WRAP if col == 3 else _CENTER
                cell.border = _THIN
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── dev_tasks.xlsx ───────────────────────────────────────────────────────

def dev_tasks_xlsx(project: str) -> bytes:
    """Dev task breakdown export, built from estimate.json (epics[].tasks[].dev_tasks[]).

    E2E rows are skipped: they have no breakdown. A task with an empty dev_tasks list
    still produces ONE row — Dev Task = the task name, empty description, Days =
    task.days — so the document handed to the developers stays complete even for
    tasks that have not been broken down yet.
    """
    data = _require(project, "estimate")
    epics = data["epics"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Dev Tasks"
    widths = [28, 40, 45, 70, 14, 40]
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w

    headers = ["Epic", "Task (Estimate)", "Dev Task", "Description for the Developer",
               "Days", "Owner"]
    _header_row(ws, headers)

    epic_groups = []
    task_groups = []
    row = 2

    for epic in epics:
        epic_start = row
        for t in epic["tasks"]:
            task_start = row
            dev_tasks = t["dev_tasks"] or [
                {"dev_task": t["task"], "description": "", "days": t["days"], "owner": ""}
            ]
            for dt in dev_tasks:
                ws.cell(row=row, column=3, value=dt["dev_task"])
                ws.cell(row=row, column=4, value=dt["description"])
                ws.cell(row=row, column=5, value=dt["days"])
                ws.cell(row=row, column=6, value=dt.get("owner") or "")

                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = _THIN
                    cell.font = _NORMAL_FONT
                    if col == 5:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                row += 1
            task_groups.append((task_start, row - 1, t["task"]))
        if row == epic_start:
            # Same failure mode as in estimate_xlsx: an epic with no tasks at all would
            # leave `end < start`, the merge would be skipped, and the epic's name would
            # land on the row the NEXT epic is about to claim as its own — overwritten and
            # gone from the file. A placeholder row keeps it visible instead.
            ws.cell(row=row, column=3, value="(no tasks yet)")
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = _THIN
                cell.font = _NORMAL_FONT
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            row += 1
        epic_groups.append((epic_start, row - 1, epic["name"]))

    for start, end, value in epic_groups:
        if start < end:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        cell = ws.cell(row=start, column=1, value=value)
        cell.font = _EPIC_FONT
        cell.fill = _EPIC_FILL
        cell.alignment = _LEFT_WRAP
        cell.border = _THIN
        for r in range(start, end + 1):
            c = ws.cell(row=r, column=1)
            c.fill = _EPIC_FILL
            c.border = _THIN

    for start, end, value in task_groups:
        if start < end:
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
        cell = ws.cell(row=start, column=2, value=value)
        cell.font = _TASK_FONT
        cell.fill = _TASK_FILL
        cell.alignment = _LEFT_WRAP
        cell.border = _THIN
        for r in range(start, end + 1):
            c = ws.cell(row=r, column=2)
            c.fill = _TASK_FILL
            c.border = _THIN

    if row > 2:
        # 6 columns (A-F: Epic, Task, Dev Task, Description, Days, Owner) — G is one past
        # the last one and used to leave a blank column inside the filter range.
        ws.auto_filter.ref = f"A1:F{row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── timeline.xlsx ────────────────────────────────────────────────────────

def timeline_xlsx(project: str) -> bytes:
    """Project plan in Excel: one sheet with the items ordered by start date, one with
    the per-developer summary.

    The dates are not stored in the JSON: they are recomputed here from the lanes in
    timeline.json (see server/lanes.py), exactly the way the board does it.
    """
    config = _require(project, "timeline")
    estimate = _require(project, "estimate")
    plan = planner.plan(estimate, config)
    names = {d["id"]: d["name"] for d in config["team"]}

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"

    states = {s["dev_task_id"]: s["status"] for s in (config.get("states") or [])}
    STATUS_LABEL = {"todo": "To do", "wip": "In progress", "done": "Done"}

    headers = ["Developer", "Status", "Epic", "Task", "Dev Task", "Layer", "Days",
               "Working days", "Start", "End"]
    _header_row(ws, headers)
    for col, w in zip("ABCDEFGHIJ", (20, 10, 30, 30, 40, 20, 8, 8, 12, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    row = 2
    for b in sorted(plan.bars, key=lambda x: (x.start, x.dev, x.item.id)):
        v = b.item
        values = [
            names.get(b.dev, b.dev),
            STATUS_LABEL[states.get(v.id, "todo")],
            v.epic_name, v.task_name or "", v.name,
            planner.LAYER_NAMES.get(v.layer, str(v.layer)),
            _round_days(v.days), b.span_days, b.start.isoformat(), b.end.isoformat(),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _THIN
            # E2E test rows stay italic here too, same as in the estimate sheet.
            cell.font = _E2E_FONT if v.task_id is None else _NORMAL_FONT
            if v.task_id is None:
                cell.fill = _E2E_FILL
            cell.alignment = _CENTER if col in (2, 6, 7, 8, 9, 10) else _LEFT_WRAP
        row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:J{row - 1}"

    ws2 = wb.create_sheet("Summary")
    _header_row(ws2, ["Developer", "Days assigned", "Leave (days declared)"])
    for col, w in zip("ABC", (24, 18, 24)):
        ws2.column_dimensions[col].width = w

    r = 2
    for d in config["team"]:
        leave = sum(
            (date.fromisoformat(p["to"]) - date.fromisoformat(p["from"])).days + 1
            for p in (d.get("leave") or [])
        )
        # 2 decimals, not 1: days come in steps of 0.25, so a sum can land on .25/.75.
        load = round(plan.load_per_dev.get(d["id"], 0), 2)
        for col, val in enumerate([d["name"], load, leave], 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.border = _THIN
            cell.font = _NORMAL_FONT
            cell.alignment = _CENTER if col > 1 else _LEFT_WRAP
        r += 1

    for label, value in (
        ("Project start", plan.start.isoformat()),
        ("Project end", plan.end.isoformat()),
        # Both ends included: a project that starts and ends on the same day lasts 1 day.
        ("Duration (calendar days)", (plan.end - plan.start).days + 1),
        ("Planned items", len(plan.bars)),
        ("Total days planned", round(sum(plan.load_per_dev.values()), 2)),
        ("Days completed", round(sum(
            b.item.days for b in plan.bars if states.get(b.item.id) == "done"), 2)),
        ("Days in progress", round(sum(
            b.item.days for b in plan.bars if states.get(b.item.id) == "wip"), 2)),
        ("Theoretical velocity (days/day)", len(config["team"])),
    ):
        ws2.cell(row=r + 1, column=1, value=label).font = _TASK_FONT
        ws2.cell(row=r + 1, column=2, value=value).font = _NORMAL_FONT
        r += 1

    if plan.unplanned:
        ws2.cell(row=r + 2, column=1, value="Items not assigned to any lane").font = _TASK_FONT
        for i, v in enumerate(plan.unplanned, 1):
            ws2.cell(row=r + 2 + i, column=1, value=f"{v.id} — {v.name}").font = _NOTE_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── data_model.drawio ────────────────────────────────────────────────────

TBL_W = 300
COL1_W = 40
COL2_W = 170
COL3_W = 90
HDR_H = 30
ROW_H = 26
STROKE = "#666666"


def _xml(value) -> str:
    """Text safe to put inside an XML attribute.

    Every label in the diagram — table names, field names, area names, relation labels —
    comes from the document, which means from whoever wrote it. One ampersand is enough
    to make the file unopenable: a data model with an area called "Stock & Inventory",
    which is an entirely ordinary name, produced a .drawio that failed to parse and an
    export that returned an error with no explanation.
    """
    return (str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))


def data_model_drawio(project: str) -> str:
    data = _require(project, "data_model")
    title = data["meta"]["title"]
    areas = {a["id"]: a for a in data["areas"]}
    tables = data["tables"]
    relations = data["relations"]

    lines: list[str] = []
    cid = 10
    pk_row_ids: dict[str, int] = {}
    fk_row_ids: dict[tuple[str, str], int] = {}
    table_geom: dict[str, tuple[int, int, int]] = {}  # name -> (x, y, h)

    def w(s):
        lines.append(s)

    # Automatic layout: a 3-column grid, tables in the order they appear in the JSON
    cols = 3
    x_gap, y_gap = 40, 60
    col_x = [20 + i * (TBL_W + x_gap) for i in range(cols)]
    col_y = [40] * cols  # next free y per column

    for idx, tbl in enumerate(tables):
        tid_ref = tbl["id"]  # reference used by relations[].from/to ("table.field")
        name = tbl["name"]   # displayed label
        area = areas.get(tbl["area"])
        color = area["color"] if area else "#4472C4"
        fields = tbl["fields"]
        n = len(fields)
        h = HDR_H + n * ROW_H

        # When the table carries a manual position (set by dragging the box in the
        # frontend editor) it is used as is — otherwise we fall back to the grid layout.
        pos = tbl.get("pos")
        if pos:
            x, y = pos["x"], pos["y"]
        else:
            col_i = idx % cols
            x = col_x[col_i]
            y = col_y[col_i]
            col_y[col_i] = y + h + y_gap
        table_geom[tid_ref] = (x, y, h)

        tbl_id = cid; cid += 1
        w(f'<mxCell id="{tbl_id}" value="{_xml(name)}" '
          f'style="shape=table;startSize={HDR_H};container=1;collapsible=0;'
          f'childLayout=tableLayout;fixedRows=1;rowLines=1;fontStyle=1;align=center;'
          f'resizeLast=1;html=1;fillColor={color};strokeColor={STROKE};fontColor=#FFFFFF;'
          f'fontSize=12;" vertex="1" parent="1">')
        w(f'  <mxGeometry x="{x}" y="{y}" width="{TBL_W}" height="{h}" as="geometry"/>')
        w('</mxCell>')

        for i, f in enumerate(fields):
            row_y = HDR_H + i * ROW_H
            row_id = cid; cid += 1
            w(f'<mxCell id="{row_id}" value="" '
              f'style="shape=tableRow;horizontal=0;startSize=0;swimlaneHead=0;swimlaneBody=0;'
              f'fillColor=none;collapsible=0;dropTarget=0;points=[[0,0.5],[1,0.5]];'
              f'portConstraint=eastwest;strokeColor=inherit;fontSize=11;" '
              f'vertex="1" parent="{tbl_id}">')
            w(f'  <mxGeometry y="{row_y}" width="{TBL_W}" height="{ROW_H}" as="geometry"/>')
            w('</mxCell>')

            key = "PK" if f.get("pk") else ("FK" if f.get("fk") else "")
            if key == "PK":
                pk_row_ids[tid_ref] = row_id
            elif key == "FK":
                fk_row_ids[(tid_ref, f["name"])] = row_id

            if key:
                c1_style = ("shape=partialRectangle;connectable=0;fillColor=none;top=0;left=0;"
                             "bottom=0;right=1;overflow=hidden;whiteSpace=wrap;html=1;align=center;"
                             "verticalAlign=middle;fontSize=10;fontStyle=1;fontColor=#1F4E79;")
            else:
                c1_style = ("shape=partialRectangle;connectable=0;fillColor=none;top=0;left=0;"
                             "bottom=0;right=1;overflow=hidden;whiteSpace=wrap;html=1;align=center;"
                             "verticalAlign=middle;fontSize=10;fontColor=#999;")
            c1_id = cid; cid += 1
            w(f'<mxCell id="{c1_id}" value="{_xml(key)}" style="{c1_style}" vertex="1" parent="{row_id}">')
            w(f'  <mxGeometry width="{COL1_W}" height="{ROW_H}" as="geometry">'
              f'<mxRectangle width="{COL1_W}" height="{ROW_H}" as="alternateBounds"/></mxGeometry>')
            w('</mxCell>')

            if key:
                c2_style = ("shape=partialRectangle;connectable=0;fillColor=none;top=0;left=0;"
                             "bottom=0;right=1;overflow=hidden;whiteSpace=wrap;html=1;align=left;"
                             "verticalAlign=middle;spacingLeft=4;fontSize=11;fontStyle=5;fontColor=#1F4E79;")
            else:
                c2_style = ("shape=partialRectangle;connectable=0;fillColor=none;top=0;left=0;"
                             "bottom=0;right=1;overflow=hidden;whiteSpace=wrap;html=1;align=left;"
                             "verticalAlign=middle;spacingLeft=4;fontSize=11;fontColor=#333;")
            c2_id = cid; cid += 1
            w(f'<mxCell id="{c2_id}" value="{_xml(f["name"])}" style="{c2_style}" vertex="1" parent="{row_id}">')
            w(f'  <mxGeometry x="{COL1_W}" width="{COL2_W}" height="{ROW_H}" as="geometry">'
              f'<mxRectangle width="{COL2_W}" height="{ROW_H}" as="alternateBounds"/></mxGeometry>')
            w('</mxCell>')

            c3_style = ("shape=partialRectangle;connectable=0;fillColor=none;top=0;left=0;bottom=0;"
                         "right=0;overflow=hidden;whiteSpace=wrap;html=1;align=right;verticalAlign=middle;"
                         "spacingRight=4;fontSize=10;fontColor=#888;")
            c3_id = cid; cid += 1
            w(f'<mxCell id="{c3_id}" value="{_xml(f["type"])}" style="{c3_style}" vertex="1" parent="{row_id}">')
            w(f'  <mxGeometry x="{COL1_W + COL2_W}" width="{COL3_W}" height="{ROW_H}" as="geometry">'
              f'<mxRectangle width="{COL3_W}" height="{ROW_H}" as="alternateBounds"/></mxGeometry>')
            w('</mxCell>')

    # FK arrows: relations[].from = "table.field", to = "table.field". The schema does not
    # constrain the string to contain a dot (unlike the HTML export's sibling loop below,
    # which only ever needs the table half and reads it with `.split(".", 1)[0]`), so a
    # relation written as "Orders" instead of "Orders.id" must not blow up the whole
    # export: skip it, the same way a relation pointing at a row we cannot find already is.
    for rel in relations:
        src_parts = rel["from"].split(".", 1)
        tgt_parts = rel["to"].split(".", 1)
        if len(src_parts) < 2 or len(tgt_parts) < 2:
            continue
        src_table, src_field = src_parts
        tgt_table, _tgt_field = tgt_parts
        src_row = fk_row_ids.get((src_table, src_field))
        tgt_row = pk_row_ids.get(tgt_table)
        if src_row is None or tgt_row is None:
            continue
        e_id = cid; cid += 1
        w(f'<mxCell id="{e_id}" value="{_xml(rel.get("label") or "")}" '
          f'style="edgeStyle=entityRelationEdgeStyle;html=1;endArrow=ERone;startArrow=ERmany;'
          f'strokeColor=#555;strokeWidth=1.5;" edge="1" source="{src_row}" target="{tgt_row}" parent="1">')
        w('  <mxGeometry relative="1" as="geometry"/>')
        w('</mxCell>')

    # Legend
    if areas:
        max_y = max((y + h for x, y, h in table_geom.values()), default=100)
        leg_y = max_y + 60
        for i, area in enumerate(areas.values()):
            lx = 20 + i * 270
            lid = cid; cid += 1
            w(f'<mxCell id="{lid}" value="" style="rounded=0;fillColor={area["color"]};'
              f'strokeColor={STROKE};" vertex="1" parent="1">')
            w(f'  <mxGeometry x="{lx}" y="{leg_y}" width="20" height="16" as="geometry"/>')
            w('</mxCell>')
            lid2 = cid; cid += 1
            w(f'<mxCell id="{lid2}" value="{_xml(area["name"])}" '
              f'style="text;html=1;align=left;verticalAlign=middle;fillColor=none;strokeColor=none;'
              f'fontSize=11;fontColor=#333;" vertex="1" parent="1">')
            w(f'  <mxGeometry x="{lx + 26}" y="{leg_y - 2}" width="220" height="20" as="geometry"/>')
            w('</mxCell>')

    tid = cid; cid += 1
    w(f'<mxCell id="{tid}" value="Data Model — {_xml(title)} ({len(tables)} tables)" '
      f'style="text;html=1;align=center;verticalAlign=middle;fillColor=none;strokeColor=none;'
      f'fontSize=18;fontStyle=1;fontColor=#333;" vertex="1" parent="1">')
    w('  <mxGeometry x="150" y="0" width="750" height="30" as="geometry"/>')
    w('</mxCell>')

    xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += '<mxfile host="app.diagrams.net">\n'
    xml += f'<diagram id="dm" name="Data Model {_xml(title)}">\n'
    xml += ('<mxGraphModel dx="1400" dy="1200" grid="1" gridSize="10" guides="1" tooltips="1" '
            'connect="1" arrows="1" fold="1" page="0" pageScale="1" pageWidth="1200" pageHeight="2300">\n')
    xml += '<root>\n<mxCell id="0"/>\n<mxCell id="1" parent="0"/>\n'
    for line in lines:
        xml += line + '\n'
    xml += '</root>\n</mxGraphModel>\n</diagram>\n</mxfile>\n'

    ET.fromstring(xml)  # validates
    return xml


# ── data_model.html ──────────────────────────────────────────────────────

def _esc(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def data_model_html(project: str) -> str:
    data = _require(project, "data_model")
    title = data["meta"]["title"]
    areas = {a["id"]: a for a in data["areas"]}
    tables = data["tables"]
    relations = data["relations"]

    # Grid layout: 3 columns, box height driven by the number of fields.
    cols = 3
    box_w = 260
    row_h = 20
    hdr_h = 34
    gap_x, gap_y = 50, 50
    margin = 40

    col_x = [margin + i * (box_w + gap_x) for i in range(cols)]
    col_y = [margin + 40] * cols  # room for the title
    positions: dict[str, tuple[int, int, int]] = {}  # table id -> x, y, h (relations[] point at the ids)

    for idx, tbl in enumerate(tables):
        tid_ref = tbl["id"]
        h = hdr_h + len(tbl["fields"]) * row_h
        # When the table carries a manual position (dragged in the frontend editor) it is
        # used as is — otherwise we fall back to the grid layout.
        pos = tbl.get("pos")
        if pos:
            x, y = pos["x"], pos["y"]
        else:
            col_i = idx % cols
            x = col_x[col_i]
            y = col_y[col_i]
            col_y[col_i] = y + h + gap_y
        positions[tid_ref] = (x, y, h)

    auto_w = margin * 2 + cols * box_w + (cols - 1) * gap_x
    # col_y always has `cols` entries (a fixed 3-column grid, never derived from the table
    # count), so it is never empty — no fallback height is ever actually needed.
    auto_h = max(col_y) + margin
    # If a manually positioned table falls outside the automatic grid, widen the canvas so
    # it still fits.
    svg_w = max([auto_w] + [x + box_w + margin for x, y, h in positions.values()])
    svg_h = max([auto_h] + [y + h + margin for x, y, h in positions.values()])

    svg_parts = []
    svg_parts.append("""<defs>
    <marker id="arrow" viewBox="0 0 10 10" refX="9" refY="5"
            markerWidth="7" markerHeight="7" orient="auto-start-reverse">
      <path d="M 0 1 L 10 5 L 0 9 z" fill="#64748b"/>
    </marker>
  </defs>""")
    svg_parts.append('<rect width="100%" height="100%" fill="#0f172a" rx="12"/>')
    svg_parts.append(
        f'<text x="{svg_w / 2}" y="30" text-anchor="middle" font-size="17" fill="white" '
        f'font-weight="bold" font-family="Arial">Data Model — {_esc(title)}</text>'
    )

    # Relations (lines between the centers of the tables involved)
    for rel in relations:
        src_table = rel["from"].split(".", 1)[0]
        tgt_table = rel["to"].split(".", 1)[0]
        if src_table not in positions or tgt_table not in positions:
            continue
        x1, y1, h1 = positions[src_table]
        x2, y2, h2 = positions[tgt_table]
        cx1, cy1 = x1 + box_w / 2, y1 + h1 / 2
        cx2, cy2 = x2 + box_w / 2, y2 + h2 / 2
        mx, my = (cx1 + cx2) / 2, (cy1 + cy2) / 2
        svg_parts.append(
            f'<path d="M {cx1} {cy1} C {mx} {cy1}, {mx} {cy2}, {cx2} {cy2}" '
            f'fill="none" stroke="#475569" stroke-width="1.6" marker-end="url(#arrow)"/>'
        )
        label = rel.get("type", "")
        svg_parts.append(f'<rect x="{mx - 16}" y="{my - 9}" width="32" height="14" rx="3" '
                          f'fill="#0f172a" opacity="0.85"/>')
        svg_parts.append(f'<text x="{mx}" y="{my + 2}" text-anchor="middle" font-size="9" '
                          f'fill="#94a3b8" font-family="Arial">{_esc(label)}</text>')

    # Tables
    for tbl in tables:
        name = tbl["name"]
        if tbl["id"] not in positions:
            continue
        x, y, h = positions[tbl["id"]]
        area = areas.get(tbl["area"])
        color = area["color"] if area else "#3b82f6"

        svg_parts.append(f'<rect x="{x + 2}" y="{y + 2}" width="{box_w}" height="{h}" rx="8" '
                          f'fill="#000" opacity="0.3"/>')
        svg_parts.append(f'<rect x="{x}" y="{y}" width="{box_w}" height="{h}" rx="8" '
                          f'fill="#111827" stroke="{color}" stroke-width="2"/>')
        svg_parts.append(f'<rect x="{x}" y="{y}" width="{box_w}" height="{hdr_h}" rx="8" fill="{color}"/>')
        svg_parts.append(f'<rect x="{x}" y="{y + hdr_h - 8}" width="{box_w}" height="8" fill="{color}"/>')
        svg_parts.append(
            f'<text x="{x + box_w / 2}" y="{y + hdr_h / 2 + 5}" text-anchor="middle" '
            f'font-size="13" fill="white" font-weight="bold" font-family="Arial">{_esc(name)}</text>'
        )
        for i, f in enumerate(tbl["fields"]):
            fy = y + hdr_h + i * row_h + row_h / 2 + 4
            badge = "PK" if f.get("pk") else ("FK" if f.get("fk") else "")
            badge_color = "#facc15" if badge == "PK" else ("#38bdf8" if badge == "FK" else "#475569")
            if badge:
                svg_parts.append(
                    f'<text x="{x + 12}" y="{fy}" font-size="9" fill="{badge_color}" '
                    f'font-weight="bold" font-family="Arial">{badge}</text>'
                )
            svg_parts.append(
                f'<text x="{x + 40}" y="{fy}" font-size="11" fill="#e2e8f0" font-family="Arial">'
                f'{_esc(f["name"])}</text>'
            )
            svg_parts.append(
                f'<text x="{x + box_w - 10}" y="{fy}" text-anchor="end" font-size="10" '
                f'fill="#94a3b8" font-family="Arial">{_esc(f["type"])}</text>'
            )

    # Area legend
    legend_y = svg_h - 20
    lx = margin
    for area in areas.values():
        svg_parts.append(f'<rect x="{lx}" y="{legend_y}" width="14" height="14" rx="3" '
                          f'fill="{area["color"]}"/>')
        svg_parts.append(f'<text x="{lx + 20}" y="{legend_y + 11}" font-size="11" fill="#94a3b8" '
                          f'font-family="Arial">{_esc(area["name"])}</text>')
        lx += 30 + len(area["name"]) * 7 + 20

    svg_content = (
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {svg_w} {svg_h}" '
        f'width="{svg_w}" height="{svg_h}">\n' + '\n'.join(svg_parts) + '\n</svg>'
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Data Model — {_esc(title)}</title>
  <style>
    body {{ margin: 0; background: #0f172a; display: flex; justify-content: center;
            align-items: flex-start; min-height: 100vh; font-family: Arial, sans-serif; }}
    .container {{ padding: 20px; max-width: 100%; overflow-x: auto; }}
  </style>
</head>
<body>
  <div class="container">
    {svg_content}
  </div>
</body>
</html>"""


# ── test_plan.xlsx ───────────────────────────────────────────────────────

# Case-level columns (identical for every step, merged vertically) vs step-level
# columns (one row per step) — same principle as the epic/task merge in
# estimate_xlsx and dev_tasks_xlsx, applied to the case instead of the epic.
_TP_CASE_COLS = (1, 2, 3, 4, 5, 6, 10, 11, 12, 13, 14)
_TP_STEP_COLS = (7, 8, 9)
_TP_CENTER_COLS = {1, 3, 4, 5, 11, 12, 13}


def test_plan_xlsx(project: str) -> bytes:
    """Sheet 'Test plan' (one row per step of each case, case columns merged vertically)
    plus sheet 'Coverage' (cases per epic of estimate.json, when available — silently
    skipped if estimate.json does not exist yet: the test plan can be written before the
    estimate is confirmed).
    """
    data = _require(project, "test_plan")
    cases = data["cases"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Test plan"

    headers = ["Case ID", "Title", "Epic", "Task", "Type", "Preconditions",
               "Step no.", "Action", "Expected result of the step",
               "Overall expected result", "Outcome", "Tester",
               "Run date", "Notes"]
    _header_row(ws, headers)
    for col, w in zip("ABCDEFGHIJKLMN",
                       [10, 30, 8, 10, 12, 28, 9, 42, 34, 34, 12, 14, 14, 30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    row = 2
    case_groups = []  # (start, end, case)

    for case in cases:
        start = row
        for step in case["steps"]:
            ws.cell(row=row, column=7, value=step["n"])
            ws.cell(row=row, column=8, value=step["action"])
            ws.cell(row=row, column=9, value=step.get("expected") or "")
            for col in _TP_STEP_COLS:
                cell = ws.cell(row=row, column=col)
                cell.font = _NORMAL_FONT
                cell.border = _THIN
                cell.alignment = _CENTER if col == 7 else _LEFT_TOP
            row += 1
        end = row - 1
        case_groups.append((start, end, case))

    for start, end, case in case_groups:
        outcome = case.get("outcome") or "to_run"
        fill = _OUTCOME_FILLS.get(outcome, _OUTCOME_NEUTRAL_FILL)

        if start < end:
            for col in _TP_CASE_COLS:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)

        vals = {
            1: case["id"], 2: case["title"], 3: case["epic"],
            4: case.get("task") or "", 5: case.get("type") or "",
            6: case.get("preconditions") or "",
            10: case.get("expected_result") or "",
            11: _OUTCOME_LABEL.get(outcome, outcome),
            12: case.get("tester") or "",
            13: case.get("run_at") or "",
            14: case.get("notes") or "",
        }
        for col, val in vals.items():
            cell = ws.cell(row=start, column=col, value=val)
            cell.font = _NORMAL_FONT
            cell.alignment = _CENTER if col in _TP_CENTER_COLS else _LEFT_TOP

        # Border over the whole merged row plus the outcome color on every row it spans
        # (needed because the non-anchor cells of a merge are MergedCell: they accept
        # fill/border but not value assignment).
        for r in range(start, end + 1):
            for col in _TP_CASE_COLS:
                c = ws.cell(row=r, column=col)
                c.border = _THIN
                if col == 11:
                    c.fill = fill

    if row > 2:
        ws.auto_filter.ref = f"A1:N{row - 1}"

    # Coverage sheet: depends on estimate.json, which is optional for the test plan
    # (it can exist before the estimate has been written) — we just skip it.
    estimate = load_doc(project, "estimate")
    if estimate:
        ws2 = wb.create_sheet("Coverage")
        _header_row(ws2, ["Epic", "Epic Name", "No. Cases", "OK", "KO", "Blocked", "Pending"])
        for col, w in zip("ABCDEFG", [10, 45, 10, 8, 8, 10, 12]):
            ws2.column_dimensions[col].width = w
        ws2.freeze_panes = "A2"

        counts = defaultdict(lambda: {"ok": 0, "ko": 0, "blocked": 0, "to_run": 0, "total": 0})
        for case in cases:
            c = counts[case["epic"]]
            c["total"] += 1
            outcome = case.get("outcome") or "to_run"
            c[outcome] += 1

        r = 2
        empty = {"ok": 0, "ko": 0, "blocked": 0, "to_run": 0, "total": 0}
        for epic in estimate["epics"]:
            c = counts.get(epic["id"], empty)
            # An epic with zero cases is the whole reason this sheet exists: it has to be
            # highlighted, not merely listed with a zero that is easy to miss.
            uncovered = c["total"] == 0
            vals = [epic["id"], epic["name"], c["total"], c["ok"], c["ko"], c["blocked"], c["to_run"]]
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=col, value=val)
                cell.border = _THIN
                cell.alignment = _LEFT_WRAP if col == 2 else _CENTER
                if uncovered:
                    cell.fill = _OUTCOME_KO_FILL
                    cell.font = Font(name="Arial", bold=True, size=10, color="9C0006")
                else:
                    cell.font = _NORMAL_FONT
            r += 1
        if r > 2:
            ws2.auto_filter.ref = f"A1:G{r - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
