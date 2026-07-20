"""Dyla backend — FastAPI + Claude Agent SDK.

Run with:  python -m uvicorn server.main:app --host 127.0.0.1 --port 3000
(start.ps1 / start.sh do exactly this — 3000 is the port they check and open a
browser on, see server/config.yaml).
"""
import re
import shutil
import threading
import uuid
from pathlib import Path

import jsonschema
from fastapi import (BackgroundTasks, Body, FastAPI, Form, HTTPException, UploadFile,
                     WebSocket, WebSocketDisconnect)
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from .config import PROJECTS_DIR, ROOT, RUNTIME_DIR
from . import agenda
from . import deck_export
from . import engine_setup
from . import dictation as dictation_mod
from . import docx_export
from . import exports
from . import ingest
from . import meetings
from . import mockup_export
from . import models
from . import preferences
from . import project_meta
from . import transcription
from .documents import DOC_NAMES, DocumentUnreadable, load_doc, save_doc
from .engine_metrics import EngineMetrics
from .model_router import ModelRouter
from .session_manager import (GLOBAL_CHAT, SessionManager, create_chat, delete_chat,
                               list_chats, load_chat, project_cost, project_totals,
                               rename_chat, set_active_chat)
from .summarize import summarize_brief
from .versioning import list_versions, restore

app = FastAPI(title="Dyla")
router = ModelRouter()
sessions = SessionManager(router)
# Holds the previous counter reading between polls, which is what makes the tokens/s
# figure a measurement rather than a stale gauge — see engine_metrics.
_engine_metrics = EngineMetrics()

NAME_RE = re.compile(r"^[a-z0-9][a-z0-9_-]{0,63}$")

# Workflow deliverables (see CLAUDE.md). "Living documents" architecture: the
# skills write JSON, and the xlsx/drawio/html files are exports generated on
# demand. The brief is not in this list: it can be an input file with any
# extension (brief.pdf, brief.docx...) or the deliverable brief.json, so whether
# a project has one is worked out separately.
DELIVERABLES = ["context.md", "estimate.json", "data_model.json", "mockup.json",
                "timeline.json", "questions.json", "people.json", "test_plan.json",
                "deck.json"]


def _project_dir(name: str, must_exist: bool = True):
    if not NAME_RE.match(name):
        raise HTTPException(400, "invalid project name (lowercase letters, digits, - and _ only)")
    d = PROJECTS_DIR / name
    if must_exist and not d.is_dir():
        raise HTTPException(404, f"project '{name}' not found")
    return d


def _workflow_status(name: str) -> dict:
    d = PROJECTS_DIR / name
    status = {f: (d / f).exists() for f in DELIVERABLES}
    # "brief" is true both for an input brief (in any format) and for the brief we
    # write ourselves: all the frontend cares about is whether the project has one.
    status["brief"] = bool(ingest.brief_file(d)) or (d / "brief.json").exists()
    return status


@app.on_event("startup")
def _startup():
    # The transcription model takes ~8 seconds to load. Doing it here, on a
    # background thread, keeps the first dictation of the day from paying that
    # cost by surprise. A failure here does not matter: the endpoint reloads the
    # model and reports the error itself.
    threading.Thread(target=transcription.preload, daemon=True).start()


@app.on_event("shutdown")
async def _shutdown():
    await sessions.close()
    # llama-server (a couple of GB once a model is loaded) used to keep running after
    # Dyla itself closed, because nothing ever held onto the Popen to terminate it.
    router.shutdown()


# --- projects ---

class ProjectIn(BaseModel):
    name: str
    client: str
    description: str | None = None
    # "brief": the brief is handed to us (input). "discovery": we write it
    # ourselves from the meetings (deliverable). See server/project_meta.py.
    source: str = project_meta.DEFAULT_SOURCE


# Placeholder for the activity, until there is a brief to derive it from.
ACTIVITY_TBD = "to be defined (waiting for the brief)"
ACTIVITY_TBD_DISCOVERY = "to be defined (waiting for the first meetings)"


def _write_context(d, name: str, client: str, description: str | None,
                   source: str) -> None:
    from datetime import date
    tbd = ACTIVITY_TBD_DISCOVERY if source == "discovery" else ACTIVITY_TBD
    source_label = ("discovery (we write the brief ourselves, from the meetings)"
                    if source == "discovery" else "brief provided (input document)")
    (d / "context.md").write_text(
        f"# Context — {name}\n\n## Setup\n"
        f"- Client: {client}\n"
        f"- Source: {source_label}\n"
        f"- Activity: {description or tbd}\n"
        f"- Created: {date.today().isoformat()}\n",
        encoding="utf-8")


def _set_activity(d, description: str) -> bool:
    """Replace the activity placeholder. Never touches an activity the user has
    already written: theirs always beats the model's summary."""
    f = d / "context.md"
    if not f.exists():
        return False
    text = f.read_text(encoding="utf-8")
    for tbd in (ACTIVITY_TBD, ACTIVITY_TBD_DISCOVERY):
        if tbd in text:
            f.write_text(text.replace(tbd, description, 1), encoding="utf-8")
            return True
    return False


def _last_modified(d) -> int:
    """When this project was last worked on: the most recent timestamp among the
    deliverables and the context. The home page uses it to float the live
    projects to the top."""
    ts = [(d / f).stat().st_mtime for f in DELIVERABLES if (d / f).exists()]
    brief = ingest.brief_file(d)
    if brief:
        ts.append((d / brief).stat().st_mtime)
    return int(max(ts)) if ts else int(d.stat().st_mtime)


@app.get("/api/projects")
def list_projects():
    PROJECTS_DIR.mkdir(exist_ok=True)
    return [{"name": d.name, "workflow": _workflow_status(d.name),
             "source": project_meta.source(d), "client": project_meta.client(d),
             "modified": _last_modified(d)}
            for d in sorted(PROJECTS_DIR.iterdir()) if d.is_dir()]


@app.post("/api/projects", status_code=201)
def create_project(body: ProjectIn):
    if not body.client.strip():
        raise HTTPException(400, "the client is required")
    if body.source not in project_meta.SOURCES:
        raise HTTPException(400, f"source must be one of {list(project_meta.SOURCES)}")
    d = _project_dir(body.name, must_exist=False)
    if d.exists():
        raise HTTPException(409, "a project with this name already exists")
    (d / "docs").mkdir(parents=True)
    if body.source == "discovery":
        # Meeting transcripts are the primary input of these projects: the folder
        # exists from the start, so it is obvious where they go.
        (d / "meetings").mkdir()
    project_meta.create(d, body.name, body.client.strip(), body.source)
    _write_context(d, body.name, body.client.strip(),
                   (body.description or "").strip() or None, body.source)
    return {"name": body.name, "workflow": _workflow_status(body.name),
            "source": body.source}


@app.get("/api/projects/{name}")
def get_project(name: str):
    d = _project_dir(name)
    totals = project_totals(name)
    meta = project_meta.load(d)
    return {"name": name, "workflow": _workflow_status(name),
            "source": meta.get("source"), "client": meta.get("client"),
            "cost_usd": totals["cost_usd"], "tokens": totals["tokens"]}


# --- chat: multi-chat registry, history and interruption (name can be a project or "_global") ---

def _chat_target(name: str) -> str:
    if name == GLOBAL_CHAT:
        return name
    _project_dir(name)
    return name


def _resolve_chat_id(name: str, chat_id: str | None) -> str:
    """The explicit chat_id when given, otherwise the project's active chat — a handy fallback
    for callers that do not pass it yet (bootstrap, older clients)."""
    if chat_id:
        return chat_id
    return list_chats(name)["active"]


class ChatIn(BaseModel):
    title: str | None = None


class ChatPatchIn(BaseModel):
    title: str | None = None
    active: bool | None = None


@app.get("/api/projects/{name}/chats")
def get_chats(name: str):
    return list_chats(_chat_target(name))


@app.post("/api/projects/{name}/chats", status_code=201)
def post_chat(name: str, body: ChatIn = Body(default=ChatIn())):
    return create_chat(_chat_target(name), body.title)


@app.patch("/api/projects/{name}/chats/{chat_id}")
def patch_chat(name: str, chat_id: str, body: ChatPatchIn):
    target = _chat_target(name)
    try:
        chat = None
        if body.title is not None:
            chat = rename_chat(target, chat_id, body.title)
        if body.active:
            chat = set_active_chat(target, chat_id)
        if chat is None:
            chat = _find_chat_or_404(target, chat_id)
    except KeyError:
        raise HTTPException(404, f"chat '{chat_id}' not found")
    return chat


@app.delete("/api/projects/{name}/chats/{chat_id}")
async def delete_chat_endpoint(name: str, chat_id: str):
    target = _chat_target(name)
    await sessions.close(target, chat_id)  # close the in-memory SDK session before deleting files
    try:
        delete_chat(target, chat_id)
    except KeyError:
        raise HTTPException(404, f"chat '{chat_id}' not found")
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True}


def _find_chat_or_404(name: str, chat_id: str) -> dict:
    for c in list_chats(name)["chats"]:
        if c["id"] == chat_id:
            return c
    raise KeyError(chat_id)


@app.get("/api/chats/{name}/history")
def chat_history(name: str, chat_id: str | None = None):
    target = _chat_target(name)
    cid = _resolve_chat_id(target, chat_id)
    return {"turns": load_chat(target, cid), "cost_usd": project_cost(target), "chat_id": cid}


@app.post("/api/chats/{name}/interrupt")
async def chat_interrupt(name: str, chat_id: str | None = None):
    target = _chat_target(name)
    cid = _resolve_chat_id(target, chat_id)
    sent = await sessions.interrupt(target, cid)
    return {"interrupted": sent}


# --- documents and deliverables ---

TARGETS = ("brief", "docs", "meetings")


def _safe_name(filename: str | None) -> str:
    name = re.sub(r"[^\w.\- ]", "_", filename or "document")
    # "." and ".." are made entirely of characters the regex above already allows through,
    # so an upload named ".." used to survive sanitisation unchanged and land the file on
    # the parent directory instead of inside it (a 500, not a controlled 400).
    if name in (".", ".."):
        return "document"
    return name


@app.post("/api/projects/{name}/documents", status_code=201)
async def upload_documents(name: str, files: list[UploadFile], bg: BackgroundTasks,
                           target: str = "docs"):
    """Upload one or more documents. Binary formats we know how to read (pdf, docx,
    xlsx, pptx) also get a markdown extract under `.extracted/`: that is what the
    agent reads and quotes, so it never has to open the binary."""
    d = _project_dir(name)
    if target not in TARGETS:
        raise HTTPException(400, f"target must be one of {list(TARGETS)}")
    if not files:
        raise HTTPException(400, "no file uploaded")
    if target == "brief" and len(files) > 1:
        raise HTTPException(400, "the brief is a single document")

    saved: list[str] = []
    for file in files:
        if target == "brief":
            # The brief keeps its own extension: a PDF saved as .md is unreadable
            # to everyone. The other brief.* files go away — there is only one.
            ext = Path(_safe_name(file.filename)).suffix.lower() or ".md"
            for old in ingest.BRIEF_EXT_ORDER:
                if old != ext:
                    (d / f"brief{old}").unlink(missing_ok=True)
                    ingest.drop_extract(d, f"brief{old}")
            dest = d / f"brief{ext}"
        else:
            dest = d / target / _safe_name(file.filename)
            dest.parent.mkdir(exist_ok=True)
        with dest.open("wb") as out:
            shutil.copyfileobj(file.file, out)
        rel = str(dest.relative_to(d)).replace("\\", "/")
        ingest.build_extract(d, rel)
        saved.append(rel)

    if target == "brief":
        # The summary costs a few seconds of model time: it must not hold up the
        # upload. It runs after the response, and context.md updates on its own.
        bg.add_task(_activity_from_brief, d)
    return {"saved": saved}


@app.delete("/api/projects/{name}/documents/{path:path}")
def delete_document(name: str, path: str):
    """Remove an input document (and its extract). Deliverables are not deleted from
    here: those have versioning instead."""
    d = _project_dir(name)
    f = _resolve_file(name, path)
    rel = str(f.relative_to(d.resolve())).replace("\\", "/")
    if not (rel.startswith(("docs/", "meetings/")) or rel.startswith("brief.")):
        raise HTTPException(400, "only input documents can be removed")
    f.unlink()
    ingest.drop_extract(d, rel)
    return {"removed": rel}


async def _activity_from_brief(d) -> None:
    """Fill in the activity line of context.md with a summary of the brief (Haiku,
    first pages). Best effort: if the brief is unreadable or the model does not
    answer, the placeholder stays."""
    text = ingest.brief_text(d)
    if not text:
        return
    summary = await summarize_brief(text, env=router.session_kwargs()["env"])
    if summary:
        _set_activity(d, summary)


# --- transcription of meeting recordings ---
#
# This lives here rather than with the documents because it is not an upload: it is
# a long job that PRODUCES a document. Half an hour of audio does not fit in one
# HTTP request, so the client uploads, gets a job back and polls until it is ready.

@app.post("/api/projects/{name}/transcriptions", status_code=202)
async def start_transcription(name: str, audio: UploadFile,
                              title: str = Form(""), date: str = Form("")):
    """Upload a recording and queue its transcription. Returns immediately: the
    result lands in `meetings/`, where `/meeting` looks for it."""
    d = _project_dir(name)
    try:
        job = meetings.start(name, d, audio.filename, audio.file,
                             title=title, when=date)
    except meetings.MeetingError as e:
        raise HTTPException(400, str(e))
    return job.public()


@app.get("/api/projects/{name}/transcriptions")
def list_transcriptions(name: str):
    """State of the project's jobs. This is what the client polls while waiting."""
    _project_dir(name)
    return {"jobs": meetings.jobs(name),
            "model": transcription.PROFILES[meetings.PROFILE].model}


@app.post("/api/projects/{name}/transcriptions/{job_id}/cancel")
def cancel_transcription(name: str, job_id: str):
    _project_dir(name)
    try:
        return meetings.cancel(job_id, name).public()
    except meetings.MeetingError as e:
        raise HTTPException(404, str(e))


@app.post("/api/projects/{name}/transcriptions/{job_id}/confirm")
def confirm_transcription(name: str, job_id: str):
    """The transcript has been proofread: the audio is dropped, since up to this
    point it was only kept to check the doubtful passages."""
    d = _project_dir(name)
    try:
        return meetings.confirm(job_id, name, d).public()
    except meetings.MeetingError as e:
        raise HTTPException(409, str(e))


@app.delete("/api/projects/{name}/transcriptions/{job_id}")
def discard_transcription(name: str, job_id: str):
    """Throw everything away: the audio and the markdown it produced."""
    d = _project_dir(name)
    try:
        return {"removed": meetings.discard(job_id, name, d)}
    except meetings.MeetingError as e:
        raise HTTPException(404, str(e))


@app.get("/api/projects/{name}/files")
def list_files(name: str):
    d = _project_dir(name)
    return sorted(
        str(f.relative_to(d)).replace("\\", "/")
        for f in d.rglob("*")
        if f.is_file() and not f.name.startswith(".")
        and not any(p.startswith(".") for p in f.relative_to(d).parts)
    )


@app.get("/api/projects/{name}/documents")
def list_documents(name: str):
    """INPUT documents with their metadata: this is what fills the document
    dropdown. Deliverables do not show up here — they have their own tabs."""
    d = _project_dir(name)
    out = []
    brief = ingest.brief_file(d)
    roots = [(brief, "brief")] if brief else []
    for sub in ("docs", "meetings"):
        folder = d / sub
        if folder.is_dir():
            roots += [(f"{sub}/{f.name}", sub) for f in sorted(folder.iterdir())
                      if f.is_file() and not f.name.startswith(".")]
    for rel, kind in roots:
        f = d / rel
        # Documents uploaded before extraction existed have no extract: we build it
        # the first time the project is opened, once and for all.
        if ingest.needs_extract(rel) and not ingest.extract_path(d, rel).is_file():
            ingest.build_extract(d, rel)
        stat = f.stat()
        ext = f.suffix.lower()
        out.append({
            "file": rel,
            "name": f.name,
            "kind": kind,
            "ext": ext.lstrip("."),
            "size": stat.st_size,
            "modified": int(stat.st_mtime),
            # "readable" = the agent can see the text (directly or via the extract).
            "readable": ext in ingest.SUPPORTED_EXT,
            "extracted": ingest.extract_path(d, rel).is_file(),
        })
    return out


@app.get("/api/projects/{name}/brief")
def get_brief(name: str):
    """State of the brief and its citable anchors (chapters), for chat citations."""
    d = _project_dir(name)
    source = project_meta.source(d)
    file = ingest.brief_file(d)
    text = ingest.brief_text(d) if file else None
    if text is None and (d / "brief.json").is_file():
        # Brief written by us: the anchors are the JSON chapters, not a file's headings.
        # load_doc is the one that can raise DocumentUnreadable (a half-written brief.json
        # from an interrupted turn) — every other caller of it in this file is wrapped for
        # exactly that reason, this one was missed and came back as an unexplained 500.
        try:
            doc = load_doc(name, "brief") or {}
        except DocumentUnreadable as e:
            raise HTTPException(409, str(e))
        chapters = doc.get("chapters", [])
        return {"source": source, "file": "brief.json", "kind": "doc",
                "headings": [{"level": 1, "title": c.get("title", ""),
                              "slug": ingest.slugify(c.get("title", "")),
                              "id": c.get("id"), "line": None, "page": None}
                             for c in chapters]}
    return {"source": source, "file": file, "kind": "file" if file else None,
            "headings": ingest.headings(text) if text else []}


def _resolve_file(name: str, path: str):
    d = _project_dir(name)
    f = (d / path).resolve()
    if not f.is_relative_to(d.resolve()) or not f.is_file():
        raise HTTPException(404, "file not found")
    return f


@app.get("/api/projects/{name}/files/{path:path}")
def download_file(name: str, path: str, dl: bool = False):
    """dl=1 forces a download; without it the file is served inline (iframe preview)."""
    f = _resolve_file(name, path)
    return FileResponse(f, filename=f.name if dl else None)


@app.get("/api/projects/{name}/preview/{path:path}")
def preview_file(name: str, path: str):
    """HTML preview for files the browser cannot render natively (xlsx)."""
    f = _resolve_file(name, path)
    if f.suffix.lower() != ".xlsx":
        raise HTTPException(400, "preview is available for xlsx files only")
    from openpyxl import load_workbook

    try:
        wb = load_workbook(f, data_only=True)
    except Exception:
        raise HTTPException(422, "not a valid xlsx file")
    import html as _html

    parts = ["<style>body{font:13px 'Segoe UI',sans-serif;margin:16px}"
             "h2{font-weight:600;font-size:15px;margin:18px 0 8px}"
             "table{border-collapse:collapse;margin-bottom:12px}"
             "td,th{border:1px solid #ccc;padding:4px 9px;text-align:left;vertical-align:top}"
             "tr:first-child td{font-weight:600;background:#f3f0ec}</style>"]
    for ws in wb.worksheets:
        # Cell values and the sheet title come straight from whatever xlsx was uploaded:
        # an <img src=x onerror=...> in a cell used to run as script in this same-origin
        # iframe, because none of it was ever escaped before landing in the HTML preview.
        parts.append(f"<h2>{_html.escape(ws.title)}</h2><table>")
        for row in ws.iter_rows(values_only=True):
            cells = "".join(f"<td>{'' if v is None else _html.escape(str(v))}</td>" for v in row)
            parts.append(f"<tr>{cells}</tr>")
        parts.append("</table>")
    from fastapi.responses import HTMLResponse

    return HTMLResponse("".join(parts))


# --- living documents (JSON) and on-demand exports ---

EXPORT_MAP = {
    "estimate.xlsx": ("estimate", exports.estimate_xlsx,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "dev_tasks.xlsx": ("estimate", exports.dev_tasks_xlsx,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "timeline.xlsx": ("timeline", exports.timeline_xlsx,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "data_model.drawio": ("data_model", exports.data_model_drawio, "application/xml"),
    "data_model.html": ("data_model", exports.data_model_html, "text/html; charset=utf-8"),
    "mockup.html": ("mockup", mockup_export.build_mockup_html, "text/html; charset=utf-8"),
    "test_plan.xlsx": ("test_plan", exports.test_plan_xlsx,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    # Word is the format the client receives the brief in, and comments on.
    "brief.docx": ("brief", docx_export.brief_docx,
                   "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
    "deck.pptx": ("deck", deck_export.deck_pptx,
                  "application/vnd.openxmlformats-officedocument.presentationml.presentation"),
    "deck.html": ("deck", deck_export.deck_html, "text/html; charset=utf-8"),
}


@app.get("/api/projects/{name}/doc/{doc}")
def get_doc(name: str, doc: str):
    if doc not in DOC_NAMES:
        raise HTTPException(400, f"unknown document: '{doc}' (expected one of: {sorted(DOC_NAMES)})")
    _project_dir(name)
    try:
        data = load_doc(name, doc)
    except DocumentUnreadable as e:
        # 409 rather than 500: nothing is broken on our side, the file on disk is in a
        # state the user can fix, and the message says how.
        raise HTTPException(409, str(e))
    if data is None:
        raise HTTPException(404, f"{doc}.json not found for project '{name}'")
    return data


@app.put("/api/projects/{name}/doc/{doc}")
def put_doc(name: str, doc: str, body: dict = Body(...)):
    if doc not in DOC_NAMES:
        raise HTTPException(400, f"unknown document: '{doc}' (expected one of: {sorted(DOC_NAMES)})")
    _project_dir(name)
    try:
        save_doc(name, doc, body)
    except jsonschema.ValidationError as e:
        path = ".".join(str(p) for p in e.absolute_path)
        msg = f"{e.message} (at {path})" if path else e.message
        raise HTTPException(422, msg)
    return {"ok": True}


@app.get("/api/projects/{name}/export/{kind}")
def get_export(name: str, kind: str, inline: bool = False):
    """inline=1 to show the export inside the UI (iframe): with
    `Content-Disposition: attachment` the browser downloads the file instead of
    rendering it, and the preview stays a blank page."""
    if kind not in EXPORT_MAP:
        raise HTTPException(400, f"unknown export: '{kind}' (expected one of: {sorted(EXPORT_MAP)})")
    _project_dir(name)
    _doc_name, generator, content_type = EXPORT_MAP[kind]
    try:
        content = generator(name)
    except exports.DocNotFound:
        raise HTTPException(404, f"the source document needed to generate '{kind}' is missing")
    except DocumentUnreadable as e:
        # The export is generated from the JSON, so an unreadable source document surfaces
        # here too — with the same explanation rather than a failed download.
        raise HTTPException(409, str(e))
    body = content if isinstance(content, (bytes, bytearray)) else content.encode("utf-8")
    disposition = "inline" if inline else f'attachment; filename="{kind}"'
    return Response(content=body, media_type=content_type,
                    headers={"Content-Disposition": disposition})


# --- personal agenda (spans projects, belongs to none of them) ---


@app.get("/api/agenda")
def get_agenda():
    data = agenda.load()
    return {**data, "buckets": agenda.group(data),
            "transcription_ready": transcription.ready()}


@app.put("/api/agenda")
def put_agenda(body: dict = Body(...)):
    try:
        agenda.save(body)
    except jsonschema.ValidationError as e:
        path = ".".join(str(p) for p in e.absolute_path)
        raise HTTPException(422, f"{e.message} (at {path})" if path else e.message)
    return {"ok": True}


class ItemIn(BaseModel):
    text: str
    projects: list[str] | None = None
    due: str | None = None
    time: str | None = None
    priority: str | None = None
    notes: str | None = None
    source: str = "manual"


@app.post("/api/agenda/items", status_code=201)
def add_items(body: list[ItemIn]):
    """Add one or more items. This is the endpoint that commits the suggestions
    coming from a dictation, once the user has reviewed them."""
    from datetime import date as _date
    data = agenda.load()
    created = _date.today().isoformat()
    new_items = []
    for v in body:
        text = v.text.strip()
        if not text:
            continue
        item = {"id": agenda.next_id(data["items"] + new_items), "text": text,
                "status": "open", "source": v.source, "created": created}
        if v.projects:
            item["projects"] = [p for p in (x.strip() for x in v.projects) if p]
        for field, value in (("due", v.due), ("time", v.time), ("priority", v.priority),
                             ("notes", v.notes)):
            if value:
                item[field] = value
        new_items.append(item)
    if not new_items:
        raise HTTPException(400, "no valid item to add")
    data["items"].extend(new_items)
    data["meta"]["date"] = created
    try:
        agenda.save(data)
    except jsonschema.ValidationError as e:
        raise HTTPException(422, e.message)
    return {"added": [v["id"] for v in new_items]}


@app.patch("/api/agenda/items/{item_id}")
def update_item(item_id: str, body: dict = Body(...)):
    """Update an item. Accepted fields: text, projects, due, priority, status,
    notes. Passing null on an optional field removes it (e.g. to clear the date)."""
    from datetime import date as _date
    data = agenda.load()
    item = next((v for v in data["items"] if v["id"] == item_id), None)
    if item is None:
        raise HTTPException(404, f"item '{item_id}' not found")
    for field in ("text", "projects", "due", "time", "priority", "status", "notes"):
        if field not in body:
            continue
        if body[field] is None:
            item.pop(field, None)
        else:
            item[field] = body[field]
    # The completion date is set by the backend: it is a fact, not a UI field.
    if item.get("status") == "done":
        item.setdefault("completed", _date.today().isoformat())
    else:
        item.pop("completed", None)
    data["meta"]["date"] = _date.today().isoformat()
    try:
        agenda.save(data)
    except jsonschema.ValidationError as e:
        raise HTTPException(422, e.message)
    return item


@app.delete("/api/agenda/items/{item_id}")
def delete_item(item_id: str):
    data = agenda.load()
    remaining = [v for v in data["items"] if v["id"] != item_id]
    if len(remaining) == len(data["items"]):
        raise HTTPException(404, f"item '{item_id}' not found")
    data["items"] = remaining
    agenda.save(data)
    return {"removed": item_id}


async def _suggestions_from_text(spoken: str, empty_reason: str) -> dict:
    spoken = (spoken or "").strip()
    if not spoken:
        return {"text": "", "items": [], "reason": empty_reason}
    PROJECTS_DIR.mkdir(exist_ok=True)
    projects = sorted(d.name for d in PROJECTS_DIR.iterdir() if d.is_dir())
    items = await dictation_mod.structure(spoken, projects,
                                          env=router.session_kwargs()["env"])
    return {"text": spoken, "items": items}


# Two endpoints instead of one with an optional parameter: putting an UploadFile and
# a Body in the same signature makes FastAPI read the whole request as form-data, and
# the JSON is silently ignored (the text always arrives empty).
@app.post("/api/agenda/dictation")
async def dictate(audio: UploadFile):
    """From a voice note to SUGGESTED items, not saved ones. Transcription runs
    locally; the caller shows the preview and commits with POST /api/agenda/items."""
    suffix = Path(_safe_name(audio.filename)).suffix or ".webm"
    # A fixed name here meant two dictations running at the same time (two tabs, or one
    # still in flight when a second is fired off) overwrote each other's audio mid-write.
    tmp = RUNTIME_DIR / f".dictation-{uuid.uuid4().hex}{suffix}"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    try:
        with tmp.open("wb") as out:
            shutil.copyfileobj(audio.file, out)
        try:
            spoken = transcription.transcribe(tmp, language=preferences.whisper_language())
        except transcription.TranscriptionUnavailable as e:
            raise HTTPException(503, str(e))
    finally:
        tmp.unlink(missing_ok=True)
    return await _suggestions_from_text(spoken, "I did not hear anything")


class ParseIn(BaseModel):
    text: str = ""


@app.post("/api/agenda/parse")
async def parse_text(body: ParseIn):
    """Same as dictation but starting from written text: same preview, no microphone."""
    return await _suggestions_from_text(body.text, "I did not find any activity in this text")


# --- deliverable versions ---

class RestoreIn(BaseModel):
    file: str
    v: int


@app.get("/api/projects/{name}/versions")
def get_versions(name: str):
    _project_dir(name)
    return list_versions(name)


@app.post("/api/projects/{name}/restore")
def restore_version(name: str, body: RestoreIn):
    _project_dir(name)
    try:
        restore(name, body.file, body.v)
    except FileNotFoundError:
        raise HTTPException(404, "version not found")
    return {"restored": body.file, "v": body.v}


# --- model ---

class ModelIn(BaseModel):
    profile: str


@app.get("/api/model")
def get_model():
    return {
        "active": router.active,
        # Only the profiles usable on THIS machine: the local model needs
        # llama-server and twenty-odd GB of GGUF, and someone without them should
        # not find a menu entry that cannot possibly work.
        "profiles": {k: (v or {}).get("label", k) for k, v in router.available().items()},
        "engine_running": router.engine_running(),
        # What the local model would run on here, and whether the engine is already in
        # place. The UI needs both to decide between "switch to local" and "install it".
        "accelerator": router.accelerator(),
        "engine_installed": router.engine_ready(),
        # The window the engine was loaded with, so the chat can say how full it is.
        # Only for a local profile: the cloud one has its own limit and we do not set it.
        "context": models.context_size(0) or None if router.is_local() else None,
    }


# --- settings: which local model runs, and where it comes from ---

@app.get("/api/models")
def list_models():
    """Everything that can be picked: what we suggest, what is already on disk, and
    what the user added themselves."""
    return {**models.catalog(),
            "accelerator": router.accelerator(),
            "engine_installed": router.engine_ready(),
            "hardware": engine_setup.hardware(),
            "context": models.context_size(0) or None,
            "context_choices": models.CONTEXT_CHOICES,
            "recommended_context": models.RECOMMENDED_CONTEXT,
            "context_advice": _context_advice()}


def _context_advice() -> dict | None:
    """What to try first on this machine, given the model that is actually selected.

    None when nothing is chosen yet: advice about a model you have not picked is noise."""
    cat = models.catalog()
    chosen = cat.get("active")
    if not chosen:
        return None
    entry = next((m for group in ("suggested", "found", "added")
                  for m in cat[group] if m["id"] == chosen), None)
    if not entry:
        return None
    hw = engine_setup.hardware()
    # On a GPU the model lives in VRAM; without one it lives in RAM.
    available = hw["vram_gb"] if router.accelerator() != "cpu" and hw["vram_gb"] else hw["ram_gb"]
    return {**models.suggest_context(entry.get("size_gb") or 0, available),
            "model": entry["label"]}


@app.post("/api/models/{model_id}/download")
def download_model(model_id: str):
    """Fetches a suggested model. Gigabytes, so this is always an explicit choice."""
    try:
        path = models.download(model_id)
    except models.ModelError as e:
        raise HTTPException(503, str(e))
    return {"downloaded": path}


class ModelPathIn(BaseModel):
    path: str
    label: str | None = None


@app.post("/api/models/added", status_code=201)
def add_model(body: ModelPathIn):
    """Registers a .gguf the user already has. Our suggestions are a starting point,
    not the limit of what Dyla will run."""
    try:
        return models.add_local(body.path, body.label)
    except models.ModelError as e:
        raise HTTPException(400, str(e))


@app.delete("/api/models/added/{model_id:path}")
def forget_model(model_id: str):
    """Forgets a model added by hand. The file stays: we did not put it there."""
    models.remove_added(model_id)
    return {"removed": model_id}


@app.put("/api/models/active")
def choose_model(body: dict = Body(...)):
    try:
        path = models.set_active(str(body.get("id") or ""))
    except models.ModelError as e:
        raise HTTPException(400, str(e))
    return {"active": body.get("id"), "path": path}


@app.put("/api/models/context")
def choose_context(body: dict = Body(...)):
    """How much context the engine loads with. It is the setting that decides whether
    the model starts at all, and only the user knows what else this machine is doing."""
    try:
        size = models.set_context(int(body.get("size") or 0))
    except (models.ModelError, ValueError) as e:
        raise HTTPException(400, str(e))
    return {"context": size, "restart_needed": router.engine_running()}


@app.post("/api/model/engine")
def install_engine():
    """Downloads the prebuilt llama-server for this machine.

    Building llama.cpp cannot be the first step of using Dyla, so we fetch the binary
    the project already publishes for this platform. Anyone with their own build sets
    LLAMA_SERVER and never comes through here."""
    try:
        path = router.install_engine()
    except engine_setup.EngineDownloadFailed as e:
        raise HTTPException(503, str(e))
    return {"installed": path, "accelerator": router.accelerator()}


@app.put("/api/model")
async def set_model(body: ModelIn):
    try:
        router.set_active(body.profile)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except RuntimeError as e:
        # This covers EngineUnavailable too: it is a RuntimeError already carrying
        # the right sentence.
        raise HTTPException(503, str(e))
    await sessions.close()  # sessions are recreated on the new profile (resume keeps the context)
    return get_model()


@app.get("/api/engine/metrics")
async def engine_metrics():
    """What the local engine is doing right now. Answers {"running": false} when there is
    nothing to report, so the UI can show or hide the panel without a second call.

    `holding` names the chat whose conversation is in the engine's memory. Everything
    about the context belongs to that one chat, and a reader looking at a different
    project would otherwise take those numbers for their own."""
    data = await _engine_metrics.read()
    if data.get("running") and sessions.loaded:
        project, chat_id = sessions.loaded
        data["holding"] = {"project": project, "chat_id": chat_id,
                           "title": _chat_title(project, chat_id)}
    return data


def _chat_title(project: str, chat_id: str) -> str:
    """The chat's name as the user renamed it, falling back to its id."""
    try:
        for c in list_chats(project)["chats"]:
            if c["id"] == chat_id:
                return c.get("title") or chat_id
    except (OSError, KeyError, ValueError):
        pass
    return chat_id


# --- preferences: how Dyla talks to you ---

@app.get("/api/preferences")
def get_preferences():
    return {"language": preferences.language()}


@app.put("/api/preferences")
async def set_preferences(body: dict = Body(...)):
    """Empty or missing clears the setting, and the agent goes back to answering in
    whatever language it was written to."""
    language = preferences.set_language(body.get("language"))
    # The language rides in the system prompt, which is fixed when a session is created:
    # without this the change would only reach chats started after it.
    await sessions.close()
    return {"language": language}


# --- chat ---

@app.websocket("/ws/{name}/{chat_id}")
async def chat_ws(ws: WebSocket, name: str, chat_id: str):
    valid = name == GLOBAL_CHAT or (NAME_RE.match(name) and (PROJECTS_DIR / name).is_dir())
    if not valid:
        await ws.close(code=4004)
        return
    await ws.accept()
    try:
        while True:
            data = await ws.receive_json()
            prompt = (data.get("prompt") or "").strip()
            anchor = data.get("anchor")
            if not prompt:
                await ws.send_json({"type": "error", "message": "empty prompt"})
                continue
            # The SDK turn must ALWAYS be consumed to the end, even if the client
            # drops halfway through the stream: cutting it short desynchronises the
            # session.
            connected = True
            try:
                async for event in sessions.chat(name, chat_id, prompt, anchor):
                    if connected:
                        try:
                            await ws.send_json(event)
                        except Exception:
                            connected = False
            except WebSocketDisconnect:
                raise
            except Exception as e:
                # Without this branch the turn died in silence: the exception closed
                # the socket and all the user saw was the spinner stopping. The
                # typical case is the `claude` CLI missing or not authenticated —
                # which is exactly the first message of someone who just installed
                # the app.
                if connected:
                    await ws.send_json({"type": "error", "message": _explain(e)})
                continue
            if not connected:
                break
    except WebSocketDisconnect:
        pass


def _explain(e: Exception) -> str:
    """Turn an exception into a sentence that says what to do. A traceback does not
    help someone who is just trying to chat."""
    name = type(e).__name__
    if name == "CLINotFoundError":
        return ("Claude Code is not installed, and the agent cannot start without it. "
                "Install it with  npm install -g @anthropic-ai/claude-code  and restart the app.")
    if name in ("ProcessError", "CLIConnectionError"):
        return (f"The agent failed to start ({e}). It is usually an access problem: "
                "run  claude  from a terminal, complete the login, then try again.")
    return f"{name}: {e}"


# --- frontend ---
# The built frontend in web/dist if present, otherwise fall back to web/ until the
# build exists.

_WEB_DIR = ROOT / "web" / "dist" if (ROOT / "web" / "dist").is_dir() else ROOT / "web"


class _NoCacheHtml(StaticFiles):
    """StaticFiles that refuses to let the HTML be cached.

    Build assets carry a hash in their filename and can be cached forever, but
    index.html cannot: it is the file that POINTS at the current hash. If the
    browser caches it, it keeps asking for the previous build's bundle, and after a
    deploy you see the old app with no idea why (a hard reload is needed).
    """

    async def get_response(self, path: str, scope):
        r = await super().get_response(path, scope)
        if path.endswith(".html") or path in ("", "."):
            r.headers["Cache-Control"] = "no-cache, must-revalidate"
        return r


# Mounted last: the /api and /ws routes take precedence, everything else
# (/, /assets/...) is served from the build (html=True serves index.html on /).
app.mount("/", _NoCacheHtml(directory=_WEB_DIR, html=True), name="webroot")
